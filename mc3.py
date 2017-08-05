# -*- coding: utf-8 -*-

"""Mc3 module."""

import cmd2

from linebot.models import (
    MessageEvent, TextMessage, TextSendMessage,
    SourceUser, SourceGroup, SourceRoom,
    TemplateSendMessage, ConfirmTemplate, MessageTemplateAction,
    ButtonsTemplate, URITemplateAction, PostbackTemplateAction,
    CarouselTemplate, CarouselColumn, PostbackEvent,
    StickerMessage, StickerSendMessage, LocationMessage, LocationSendMessage,
    ImageMessage, VideoMessage, AudioMessage,
    UnfollowEvent, FollowEvent, JoinEvent, LeaveEvent, BeaconEvent, ImageSendMessage,VideoSendMessage
)

from openpyxl import Workbook, load_workbook


from aq import (aq, maps, boss)
from arena import arena
#from aw import aw
from mastery import mastery
from mcoccalendars import calendar
from prestige import prestige
#from prestigetest import inputchamp
from sig import sig
from newcalc import (inputchamp, lookup,test)
from cutoffs import cutoffs

from utils import logcall

lolfights={'abomination':'https://vimeo.com/220837667/9c98696855','agentvenom':'https://vimeo.com/220838058/bd0e862d6f',
           'blackwidow':'https://vimeo.com/220838431/356057d41f', 'civilwarrior':'https://vimeo.com/220840647/f153553a84',
           'colossus':'https://vimeo.com/220841133/6657fd3a46','cyclops':'https://vimeo.com/220846104/3e21d9f084','daredevilnetflix':'https://vimeo.com/220846544/be827a3b35',
          'drstrange':'https://vimeo.com/220847138/24d98cc720','electro':'https://vimeo.com/220847623/82e3d39474','elektra':'https://vimeo.com/220848049/884435b896',
          'groot':'https://vimeo.com/220857481/c489f351d1','guillotine':'https://vimeo.com/223911603/fa7dafe181','hulkbuster':'https://vimeo.com/223911768/d9a2620060',
           'juggernaut':'https://vimeo.com/223911847/ee3e9d8617','kamalakhan':'https://vimeo.com/223911918/1d7f2ff0a7', 'maestro':'https://vimeo.com/223911941/167db375ec',
           'magneto':'https://vimeo.com/223912048/4240e32288', 'milesmorales':'https://vimeo.com/223912103/48c2abd59f','moonknight':'https://vimeo.com/223912179/a8d4d48bb7',
           'msmarvel':'https://vimeo.com/223912244/6211f8a7af','ogvision':'https://vimeo.com/223912305/48fe9b5c39',
           'spidergwen':'https://vimeo.com/223911690/1e59d7e6f6'
           
          }

champ_bios = load_workbook('champbios.xlsx') 

class Mc3(cmd2.Cmd):
    trigger = 'Mc3 '

    def __init__(self, line_bot_api, event):
        self.line_bot_api = line_bot_api
        self.event = event
        super(Mc3, self).__init__()
        
    #def default(self, line):
         #self.line_bot_api.reply_message(self.event.reply_token,
         #   TextSendMessage(text="Oops! You've entered an invalid command please type " + 
         #   self.trigger + "list to view available commands"))
        #return False
    
    def do_list(self, line):
        self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="Make sure all commands start with the 'Mc3' trigger"+'\n'+                                    
                                     "Example:'Mc3 aq'"+'\n'+
                                     '\n'+
                                     "Command List:"+'\n'+
                                     "abilities"+'\n'+
                                     "aq"+'\n'+
                                     "aw"+'\n'+
                                     "arena"+'\n'+
                                     "calendar"+'\n'+
                                     "conciergemain"+'\n'+                                    
                                     "mastery"+'\n'+
                                     "prestige"+'\n'+
                                     "prestige instructions"+'\n'+
                                     "line prestige"+'\n'+
                                     "sparring"+'\n'+
                                     "special quests"+'\n'+
                                     "Mc3 sig_calculator"+'\n'+
                                     "synergies"+'\n'+
                                     "support"+'\n'+
                                     "Bot created by Jpags, TLT, and MCOCConcierge"))
                                        
                               
        return True;

    def do_aq(self, line):
        _aq = aq(self.line_bot_api, self.event)
        
        return _aq.process(line.parsed.args)

    def do_abilities(self, line):
        
        if line == '':
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="To use Mc3 abilities, you must type: 'Mc3 abilities (champname)'. An example would be: 'Mc3 abilities spidergwen'"))
            return True

        try:
            name=line
            ws=champ_bios[name]
            signature=ws.cell(column=2,row=4).value
            passive=ws.cell(column=2,row=8).value
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="Signature Ability: "+'\n'+signature+'\n'+'\n'+
                                "Passive Ability: "+'\n'+passive))

        except KeyError:
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="Incorrect champ name, see 'Mc3 champ list' for champ name syntax"))

        return True

    def do_maps(self, line):
        _maps = maps(self.line_bot_api, self.event)
        
        return _maps.process(line.parsed.args)
    
    def do_boss(self, line):
        _boss = boss(self.line_bot_api, self.event)
        
        return _boss.process(line.parsed.args)

    def do_arena(self, line):
        _arena = arena(self.line_bot_api, self.event)
        
        return _arena.process(line.parsed.args)
    
    #def do_aw(self, line):
    #    _aw = aw(self.line_bot_api, self.event)
    #    
    #    return _aw.process(line.parsed.args)
    
    def do_calendar(self, line):
        _calendar = calendar(self.line_bot_api, self.event)
        
        return _calendar.process(line.parsed.args)
    
    def do_cutoffs(self, line):
        _cutoffs = cutoffs(self.line_bot_api, self.event)
        
        return _cutoffs.cutoffs(line)
    
    def do_inputchamp(self, line):
        _inputchamp=inputchamp(self.line_bot_api, self.event)
        
        
        return _inputchamp.calculate(line)
    
    def do_lookup(self, line):
        _lookup=lookup(self.line_bot_api, self.event)
        
        
        return _lookup.lookup(line)
    
    def do_test(self, line):
        _test=test(self.line_bot_api, self.event)
        
        
        return _test.test(line)

    
    def do_sig(self,line):
        _sig=sig(self.line_bot_api,self.event)

        return _sig.calculate(line)

    def do_lolfight(self,line):
        eventText=self.event.message.text
        trigger="Mc3 lolfight "
        opponent = eventText[eventText.find(trigger) + len(trigger):]
        video=lolfights.get(opponent, None)
        if video: 
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text=opponent+":"+video))
        else:
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text=opponent+": is not a valid lol name"))
        return True    
    
    def do_mastery(self, line):
        _mastery=mastery(self.line_bot_api, self.event)
        
        
        return _mastery.process(line.parsed.args)  

    def do_prestige(self, line):
        _prestige=prestige(self.line_bot_api, self.event)
        
        
        return _prestige.process(line.parsed.args)    
    
    def do_conciergemain(self, line):
        self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="https://line.me/R/ti/p/%40tln3406m"))
        
        
        return True
        
    def do_support(self, line):
        self.line_bot_api.reply_message(
            self.event.reply_token,
            TextSendMessage(text="To join the support room, click here: http://line.me/ti/g/LHXthtXoCV"))
        return True

    def do_help(self, line):
        self.line_bot_api.reply_message(
            self.event.reply_token,
            ImageSendMessage(
                original_content_url='https://i.imgur.com/yoY5Pyt.jpg',
                preview_image_url='https://i.imgur.com/yoY5Pyt.jpg'))
        return True

    def do_synergies(self, line):  
        self.line_bot_api.reply_message(
            self.event.reply_token,
            TextSendMessage(text="attack teams"+'\n'+"bleed teams"+'\n'+"crit teams"+'\n'+"power gain teams")) 

    def do_attack(self, line):
        if line.lower() == 'teams':
            self.line_bot_api.reply_message(
                self.event.reply_token,
                ImageSendMessage(
                    original_content_url='https://i.imgur.com/WTyyl7t.jpg',
                    preview_image_url='https://i.imgur.com/WTyyl7t.jpg'))
            return True
    
    def do_bleed(self, line):
        if line.lower() == 'teams': 
            self.line_bot_api.reply_message(
                self.event.reply_token,
                TextSendMessage(text="Coming Soon"))
            return True
            
    def do_crit(self, line):
        if line.lower() == 'teams': 
            self.line_bot_api.reply_message(
                self.event.reply_token,
                ImageSendMessage(
                    original_content_url='https://i.imgur.com/UsIqiH4.jpg',
                    preview_image_url='https://i.imgur.com/UsIqiH4.jpg')) 
            return True
            
    def do_power(self, line):
        if line.lower() == 'gain teams': 
            self.line_bot_api.reply_message(
                self.event.reply_token,
                    ImageSendMessage(
                        original_content_url='https://i.imgur.com/LBF0Rr9.jpg',
                        preview_image_url='https://i.imgur.com/LBF0Rr9.jpg'))  
            return True
            
    def do_unique(self, line):
        if line.lower() == 'teams': 
            self.line_bot_api.reply_message(
                self.event.reply_token,
                ImageSendMessage(
                    original_content_url='https://i.imgur.com/Zmec8Hs.jpg',
                    preview_image_url='https://i.imgur.com/Zmec8Hs.jpg'))   
            return True 
    
     

    #def do_my(self, line):
    #    _my=my(self.line_bot_api,self.event)

    #    return _my.process(line.parsed.args)

    #def do_help(self, line):
    #    return self.do_list(line);
        
    def do_EOF(self, line):
        return True

    @logcall
    def process(self, line):
        return self.onecmd(line);
